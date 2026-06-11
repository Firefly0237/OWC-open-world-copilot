from __future__ import annotations

import json

from openpyxl import Workbook

from owcopilot.cli.main import main
from owcopilot.content.models import ContentBundle, Entity, EntityType, Quest
from owcopilot.content.store import ContentStore


def _read_json(capsys) -> dict:
    return json.loads(capsys.readouterr().out)


def _read_stderr_json(capsys) -> dict:
    return json.loads(capsys.readouterr().err)


def test_cli_ingest_dry_run_and_write(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    source = tmp_path / "entities.json"
    source.write_text(
        json.dumps([{"kind": "entity", "id": "npc_aldric", "name": "Aldric", "type": "npc"}]),
        encoding="utf-8",
    )

    assert main(["ingest", "--content-root", str(content_root), "--input", str(source)]) == 0
    dry_run_body = _read_json(capsys)
    assert dry_run_body["dry_run"] is True
    assert dry_run_body["incoming_count"] == 1
    assert dry_run_body["cost_budget"]["used_usd"] == 0.0
    assert ContentStore(content_root).load().entities == {}

    assert (
        main(["ingest", "--content-root", str(content_root), "--input", str(source), "--write"])
        == 0
    )
    write_body = _read_json(capsys)
    assert write_body["dry_run"] is False
    assert "npc_aldric" in ContentStore(content_root).load().entities


def test_cli_ingest_uses_per_file_field_mapping_transactionally(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    faction_path = tmp_path / "阵营表.xlsx"
    relation_path = tmp_path / "阵营关系表.xlsx"
    mapping_path = tmp_path / "field_mapping.json"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["编号", "名称", "类型"])
    sheet.append(["fac_a", "A", "faction"])
    sheet.append(["fac_b", "B", "faction"])
    workbook.save(faction_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["来源阵营", "关系", "目标阵营"])
    sheet.append(["fac_a", "enemy_of", "fac_b"])
    workbook.save(relation_path)

    mapping_path.write_text(
        json.dumps(
            {
                "阵营表.xlsx": {
                    "type": "entity",
                    "columns": {"编号": "id", "名称": "name", "类型": "type"},
                },
                "阵营关系表.xlsx": {
                    "type": "relation",
                    "columns": {"来源阵营": "source", "关系": "kind", "目标阵营": "target"},
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    assert (
        main(
            [
                "ingest",
                "--content-root",
                str(content_root),
                "--field-mapping",
                str(mapping_path),
                "--input",
                str(faction_path),
                "--input",
                str(relation_path),
                "--write",
            ]
        )
        == 0
    )
    body = _read_json(capsys)
    bundle = ContentStore(content_root).load()

    assert body["incoming_count"] == 3
    assert "fac_a" in bundle.entities
    assert [(relation.source, relation.kind, relation.target) for relation in bundle.relations] == [
        ("fac_a", "enemy_of", "fac_b")
    ]


def test_cli_ingest_field_mapping_falls_back_to_matching_columns(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    task_path = tmp_path / "new_version_task.xlsx"
    mapping_path = tmp_path / "field_mapping.json"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["任务ID", "任务名", "接取NPC", "发生地点", "任务目标", "文本Key"])
    sheet.append(["q1", "New Quest", "npc_aldric", "poi_dock", "talk:npc_aldric", "QUEST_Q1"])
    workbook.save(task_path)
    mapping_path.write_text(
        json.dumps(
            {
                "02_配置表/任务表.xlsx": {
                    "type": "quest",
                    "columns": {
                        "任务ID": "id",
                        "任务名": "title",
                        "接取NPC": "giver_npc",
                        "发生地点": "location",
                        "任务目标": "objective",
                        "文本Key": "localization_keys",
                    },
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    assert (
        main(
            [
                "ingest",
                "--content-root",
                str(content_root),
                "--field-mapping",
                str(mapping_path),
                "--input",
                str(task_path),
                "--write",
            ]
        )
        == 0
    )
    _read_json(capsys)

    assert ContentStore(content_root).load().quests["q1"].objective == "talk:npc_aldric"


def test_cli_ingest_skip_conflicts_writes_non_conflicting_objects(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    source = tmp_path / "quests.json"
    ContentStore(content_root).save(
        ContentBundle(
            quests={
                "q1": Quest(
                    id="q1",
                    title="Original",
                    objective="keep existing",
                )
            }
        )
    )
    source.write_text(
        json.dumps(
            [
                {
                    "kind": "quest",
                    "id": "q1",
                    "title": "Changed",
                    "objective": "would overwrite",
                },
                {
                    "kind": "quest",
                    "id": "q2",
                    "title": "New",
                    "objective": "add this",
                },
            ]
        ),
        encoding="utf-8",
    )

    assert (
        main(
            [
                "ingest",
                "--content-root",
                str(content_root),
                "--input",
                str(source),
                "--write",
                "--skip-conflicts",
            ]
        )
        == 0
    )
    body = _read_json(capsys)
    bundle = ContentStore(content_root).load()

    assert body["has_errors"] is True
    assert body["issues"][0]["rule_code"] == "IMPORT_CONFLICT"
    assert bundle.quests["q1"].title == "Original"
    assert bundle.quests["q2"].title == "New"


def test_cli_audit_persists_and_issues_lists_filters(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    ContentStore(content_root).save(
        ContentBundle(quests={"q1": Quest(id="q1", title="Q1", giver_npc="npc_missing")})
    )

    assert main(["audit", "--content-root", str(content_root)]) == 0
    audit_body = _read_json(capsys)
    assert audit_body["open_errors"] >= 1
    assert "UNKNOWN_ENTITY_REF" in {issue["rule_code"] for issue in audit_body["issues"]}
    assert audit_body["cost_budget"]["used_usd"] == 0.0

    assert (
        main(
            [
                "issues",
                "--content-root",
                str(content_root),
                "--rule-code",
                "UNKNOWN_ENTITY_REF",
                "--status",
                "open",
            ]
        )
        == 0
    )
    issues_body = _read_json(capsys)
    assert issues_body["count"] == 1
    assert issues_body["issues"][0]["target_ref"] == "quest:q1"
    assert issues_body["cost_budget"]["used_usd"] == 0.0


def test_cli_audit_fail_on_error_output_and_baseline(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    output_path = tmp_path / "reports" / "audit.json"
    ContentStore(content_root).save(
        ContentBundle(quests={"q1": Quest(id="q1", title="Q1", giver_npc="npc_missing")})
    )

    exit_code = main(
        [
            "audit",
            "--content-root",
            str(content_root),
            "--fail-on-error",
            "--output",
            str(output_path),
        ]
    )
    audit_body = _read_json(capsys)
    assert exit_code == 1
    assert output_path.exists()
    assert json.loads(output_path.read_text(encoding="utf-8")) == audit_body

    baseline = tmp_path / "baseline.json"
    baseline.write_text(
        json.dumps({"fingerprints": [issue["fingerprint"] for issue in audit_body["issues"]]}),
        encoding="utf-8",
    )

    exit_code = main(
        [
            "audit",
            "--content-root",
            str(content_root),
            "--fail-on-error",
            "--baseline",
            str(baseline),
        ]
    )
    suppressed_body = _read_json(capsys)
    assert exit_code == 0
    assert suppressed_body["open_errors"] == 0
    assert {issue["status"] for issue in suppressed_body["issues"]} == {"suppressed"}


def test_cli_context_pack_and_ask(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    ContentStore(content_root).save(
        ContentBundle(
            entities={
                "npc_aldric": Entity(
                    id="npc_aldric",
                    name="Aldric",
                    type=EntityType.NPC,
                    description="Caravan master",
                )
            }
        )
    )

    assert (
        main(
            [
                "context-pack",
                "--content-root",
                str(content_root),
                "--query",
                "Aldric caravan",
            ]
        )
        == 0
    )
    context_body = _read_json(capsys)
    assert context_body["refs"] == ["entity:npc_aldric"]
    assert context_body["cost_budget"]["used_usd"] == 0.0

    assert (
        main(["ask", "--content-root", str(content_root), "--query", "Who is Aldric?"])
        == 0
    )
    ask_body = _read_json(capsys)
    assert ask_body["answer"]["citations"][0]["ref"] == "entity:npc_aldric"
    assert ask_body["telemetry"]["calls"] == 1
    assert ask_body["cost_budget"]["over_budget"] is False

    assert (
        main(
            [
                "ask",
                "--content-root",
                str(content_root),
                "--query",
                "Who is Aldric?",
                "--max-cost-usd",
                "0",
            ]
        )
        == 0
    )
    over_budget_body = _read_json(capsys)
    assert over_budget_body["cost_budget"]["over_budget"] is True


def test_cli_export_writes_engine_scoped_bundle(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    output_root = tmp_path / "exports"
    ContentStore(content_root).save(
        ContentBundle(
            entities={"npc_aldric": Entity(id="npc_aldric", name="Aldric", type=EntityType.NPC)}
        )
    )

    assert (
        main(
            [
                "export",
                "--content-root",
                str(content_root),
                "--output-dir",
                str(output_root),
                "--target-engine",
                "unity",
            ]
        )
        == 0
    )
    body = _read_json(capsys)
    export_dir = output_root / "unity"
    assert body["output_dir"] == str(export_dir)
    assert (export_dir / "content_bundle.json").exists()
    assert (export_dir / "manifest.json").exists()
    assert body["manifest"]["target_engine"] == "unity"
    assert body["cost_budget"]["used_usd"] == 0.0


def test_cli_ask_refuses_without_context(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    ContentStore(content_root).save(ContentBundle())

    assert (
        main(["ask", "--content-root", str(content_root), "--query", "Who is Aldric?"])
        == 0
    )
    body = _read_json(capsys)
    assert body["answer"]["refused"] is True
    assert body["telemetry"]["calls"] == 0
    assert body["cost_budget"]["used_usd"] == 0.0


def test_cli_reports_invalid_content_root(tmp_path, capsys) -> None:
    missing_root = tmp_path / "missing"

    assert main(["audit", "--content-root", str(missing_root)]) == 2
    body = _read_stderr_json(capsys)
    assert body["type"] == "FileNotFoundError"
    assert "content root does not exist" in body["error"]


def test_cli_reports_missing_input_file(tmp_path, capsys) -> None:
    content_root = tmp_path / "content"
    missing_source = tmp_path / "missing.json"

    assert (
        main(["ingest", "--content-root", str(content_root), "--input", str(missing_source)])
        == 2
    )
    body = _read_stderr_json(capsys)
    assert body["type"] == "FileNotFoundError"


def test_cli_eval_golden_runs_offline_report(tmp_path, capsys) -> None:
    workspace = tmp_path / "golden_eval"
    output_path = tmp_path / "report.json"

    assert (
        main(["eval-golden", "--workspace", str(workspace), "--output", str(output_path)])
        == 0
    )
    body = _read_json(capsys)
    assert body["passed"] is True
    assert output_path.exists()
    assert (workspace / "exports" / "generic" / "manifest.json").exists()
