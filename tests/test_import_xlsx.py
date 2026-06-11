from __future__ import annotations

from openpyxl import Workbook

from owcopilot.content.importers.xlsx import XLSXImporter
from owcopilot.content.mapping import FieldMapping, apply_field_mapping
from owcopilot.content.normalize import normalize_raw_objects


def test_xlsx_importer_reads_rows_with_source_refs(tmp_path) -> None:
    path = tmp_path / "entities.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NPCs"
    sheet.append(["kind", "id", "name", "type", "description"])
    sheet.append(["entity", "npc_aldric", "Aldric", "npc", "Caravan master"])
    workbook.save(path)

    raw = XLSXImporter().parse(path)
    bundle = normalize_raw_objects(raw)

    entity = bundle.entities["npc_aldric"]
    assert entity.description == "Caravan master"
    assert entity.source_ref is not None
    assert entity.source_ref.sheet == "NPCs"
    assert entity.source_ref.row == 2


def test_xlsx_import_can_use_field_mapping_for_chinese_headers(tmp_path) -> None:
    path = tmp_path / "entities.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["编号", "名称", "类型", "描述"])
    sheet.append(["npc_mara", "Mara", "npc", "Scout"])
    workbook.save(path)

    raw = XLSXImporter().parse(path)
    mapped = apply_field_mapping(
        raw,
        FieldMapping(
            columns={
                "编号": "id",
                "名称": "name",
                "类型": "type",
                "描述": "description",
            },
            default_kind="entity",
        ),
    )
    bundle = normalize_raw_objects(mapped)

    assert bundle.entities["npc_mara"].name == "Mara"
    assert bundle.entities["npc_mara"].description == "Scout"


def test_xlsx_importer_supports_luban_multiline_headers(tmp_path) -> None:
    path = tmp_path / "npc_luban.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TbNpc"
    sheet.append(["##var", "id", "name", "faction", "status", "desc"])
    sheet.append(["##type", "string", "string", "string", "string", "string"])
    sheet.append(["##group", "", "c,s", "c,s", "s", "c"])
    sheet.append(["##", "编号", "名称", "阵营", "状态", "描述"])
    sheet.append(["id", "name", "faction", "status", "desc"])
    sheet.append(["npc_aldric", "Aldric", "fac_guard", "active", "Caravan master"])
    workbook.save(path)

    raw = XLSXImporter().parse(path)
    bundle = normalize_raw_objects(raw)

    assert len(raw) == 1
    assert raw[0].data["id"] == "npc_aldric"
    assert bundle.entities["npc_aldric"].type.value == "npc"
    assert bundle.entities["npc_aldric"].description == "Caravan master"
    assert ("npc_aldric", "member_of", "fac_guard") in {
        (relation.source, relation.kind, relation.target) for relation in bundle.relations
    }
