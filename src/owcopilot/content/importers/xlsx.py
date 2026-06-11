"""XLSX importer for read-only spreadsheet ingestion."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from .base import RawObject


class XLSXImporter:
    def __init__(self, *, default_kind: str = "entity") -> None:
        self.default_kind = default_kind

    def parse(self, path: str | Path) -> list[RawObject]:
        from openpyxl import load_workbook

        source = Path(path)
        workbook = load_workbook(source, read_only=True, data_only=True)
        objects: list[RawObject] = []
        try:
            for sheet in workbook.worksheets:
                raw_rows = list(sheet.iter_rows(values_only=True))
                if not raw_rows:
                    continue
                headers, data_start, skip_header_echo = _table_shape(raw_rows)
                if not headers:
                    continue
                for row_number, row in enumerate(raw_rows[data_start:], start=data_start + 1):
                    if _is_comment_or_luban_meta_row(row):
                        continue
                    if skip_header_echo and _is_header_echo(headers, row):
                        continue
                    data = _row_data(headers, row)
                    if not data:
                        continue
                    kind = str(
                        data.get("kind") or data.get("object_type") or self.default_kind
                    ).lower()
                    objects.append(
                        RawObject(
                            kind=kind,
                            data=data,
                            source_path=str(source),
                            sheet=sheet.title,
                            row=row_number,
                            line=row_number,
                        )
                    )
        finally:
            workbook.close()
        return objects


def _headers(row: tuple[Any, ...]) -> list[str]:
    return [str(value).strip() if value is not None else "" for value in row]


def _table_shape(rows: list[tuple[Any, ...]]) -> tuple[list[str], int, bool]:
    """Return (headers, zero-based data_start, skip_header_echo).

    Luban tables commonly start with:
      ##var   id name ...
      ##type  ...
      ##group ...
      ##      human comments
      id      name ...
      data...

    The canonical field names are on the ##var row, shifted one column right.
    """
    first = _headers(rows[0])
    if first and first[0] == "##var":
        headers = [header for header in first[1:] if header]
        data_start = 1
        while data_start < len(rows) and _is_comment_or_luban_meta_row(rows[data_start]):
            data_start += 1
        return headers, data_start, True
    return first, 1, False


def _row_data(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = row[index] if index < len(row) else None
        if value is not None and value != "":
            data[header] = value
    return data


def _is_comment_or_luban_meta_row(row: tuple[Any, ...]) -> bool:
    first = row[0] if row else None
    return isinstance(first, str) and first.strip().startswith("##")


def _is_header_echo(headers: list[str], row: tuple[Any, ...]) -> bool:
    values = [str(value).strip() if value is not None else "" for value in row[: len(headers)]]
    return values == headers
