# ruff: noqa: E501, F401, I001
from __future__ import annotations

import hashlib
import json
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from owcopilot.assist.barks import BarkBatchService
from owcopilot.assist.drafts import QuestDraftService
from owcopilot.assist.review_queue import ReviewQueue
from owcopilot.audit.baseline import AuditBaseline, issue_fingerprint
from owcopilot.audit.context import AuditContext
from owcopilot.audit.default_rules import build_default_rule_registry
from owcopilot.audit.models import Issue
from owcopilot.audit.runner import AuditRunner
from owcopilot.content.hash import content_hash
from owcopilot.content.ingest import ingest_paths, parse_paths
from owcopilot.content.mapping import FieldMapping
from owcopilot.content.models import (
    ContentBundle,
    DialogueRef,
    Entity,
    EntityType,
    POI,
    Quest,
    RegionBrief,
    Relation,
    SourceRef,
    StyleGuide,
    Term,
)
from owcopilot.content.normalize import normalize_raw_objects
from owcopilot.content.store import ContentStore
from owcopilot.fakes import MockProvider
from owcopilot.impact.analyzer import ImpactAnalyzer
from owcopilot.impact.models import Change, ChangeSet, ChangeType
from owcopilot.llm.cache import NoOpCache
from owcopilot.llm.gateway import LLMGateway
from owcopilot.llm.router import StaticRouter
from owcopilot.llm.telemetry import TelemetryCollector
from owcopilot.patches.apply import apply_patch_to_store, rollback_patch_in_store
from owcopilot.patches.models import PatchCandidate, PatchOp, PatchOperation
from owcopilot.patches.parser import parse_patch_candidates
from owcopilot.patches.validate import valid_patch_candidates, validate_patch_candidate
from owcopilot.pipeline.project import ProjectContext


ROOT = Path(r"F:\openworld")
PY = ROOT / ".venv" / "Scripts" / "python.exe"
DATA = Path(r"F:\wuyin")
CONTENT = Path(r"F:\wuyin_content")
REPORTS = Path(r"F:\wuyin_reports")
EXPORTS = Path(r"F:\wuyin_exports")
RUN = REPORTS / datetime.now().strftime("%Y%m%d_%H%M%S")
BASELINE = CONTENT / ".owcopilot" / "baseline.json"
COMMANDS: list[dict[str, Any]] = []


def j(value: Any) -> Any:
    if hasattr(value, "model_dump"):
        return value.model_dump(mode="json")
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, set):
        return sorted(value)
    if isinstance(value, tuple):
        return [j(x) for x in value]
    if isinstance(value, list):
        return [j(x) for x in value]
    if isinstance(value, dict):
        return {str(k): j(v) for k, v in value.items()}
    return value


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(j(payload), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def reset_dir(path: Path) -> None:
    if path.resolve() not in {CONTENT.resolve(), EXPORTS.resolve()}:
        raise RuntimeError(f"refusing to reset {path}")
    path.mkdir(parents=True, exist_ok=True)
    for child in path.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()


def run_cli(name: str, args: list[str], output: Path | None = None) -> dict[str, Any]:
    cmd = [str(PY), "-m", "owcopilot.cli.main", *args]
    if output:
        cmd += ["--output", str(output)]
    t0 = time.perf_counter()
    p = subprocess.run(
        cmd,
        cwd=ROOT,
        encoding="utf-8",
        errors="replace",
        text=True,
        capture_output=True,
        check=False,
    )
    elapsed = time.perf_counter() - t0
    parsed = None
    stdout = p.stdout or ""
    stderr = p.stderr or ""
    if stdout.strip():
        try:
            parsed = json.loads(stdout.strip().splitlines()[-1])
        except json.JSONDecodeError:
            parsed = None
    record = {
        "name": name,
        "command": cmd,
        "exit_code": p.returncode,
        "duration_sec": round(elapsed, 4),
        "stdout": stdout.strip(),
        "stderr": stderr.strip(),
        "output": str(output) if output else None,
        "parsed": parsed,
    }
    COMMANDS.append(record)
    write_json(RUN / "commands" / f"{name}.json", record)
    return record


def rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(v).strip() if v is not None else "" for v in values[0]]
        out: list[dict[str, Any]] = []
        for row_no, row in enumerate(values[1:], start=2):
            item: dict[str, Any] = {"_row": row_no, "_sheet": ws.title}
            for i, header in enumerate(headers):
                if not header:
                    continue
                value = row[i] if i < len(row) else None
                if value is not None and value != "":
                    item[header] = value
            if any(not k.startswith("_") for k in item):
                out.append(item)
        return out
    finally:
        wb.close()


def split(v: Any) -> list[str]:
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    return [x.strip() for x in re.split(r"[;,；、|]", str(v).strip()) if x.strip()]


def opt_int(v: Any) -> int | None:
    if v is None or v == "" or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    text = str(v).strip()
    return int(text) if text.lstrip("-").isdigit() else None


def src(path: Path, row: dict[str, Any] | None = None) -> SourceRef:
    return SourceRef(
        path=str(path),
        sheet=str(row.get("_sheet")) if row else None,
        row=int(row["_row"]) if row else None,
    )


def safe_id(value: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "_", value).strip("_").lower()
    return text or hashlib.sha1(value.encode("utf-8")).hexdigest()[:8]


def counts(b: ContentBundle) -> dict[str, int]:
    return {
        "entities": len(b.entities),
        "relations": len(b.relations),
        "quests": len(b.quests),
        "regions": len(b.regions),
        "pois": len(b.pois),
        "dialogues": len(b.dialogues),
        "terms": len(b.terms),
        "style_guides": len(b.style_guides),
    }


def quests_from(path: Path) -> list[Quest]:
    result: list[Quest] = []
    for r in rows(path):
        result.append(
            Quest(
                id=str(r["任务ID"]),
                title=str(r["任务名"]),
                giver_npc=str(r["接取NPC"]) if r.get("接取NPC") else None,
                location=str(r["发生地点"]) if r.get("发生地点") else None,
                objective=str(r.get("任务目标") or ""),
                prerequisites=split(r.get("前置任务")),
                timeline_order=opt_int(r.get("时间线序")),
                dialogue_refs=split(r.get("对白引用")),
                localization_keys=split(r.get("文本Key")),
                tags=split(r.get("类型")),
                metadata={
                    "region_id": r.get("所属区域"),
                    "required_level": r.get("需求等级"),
                    "rewards_raw": r.get("奖励"),
                },
                source_ref=src(path, r),
            )
        )
    return result


def build_bundle() -> tuple[ContentBundle, dict[str, Any]]:
    b = ContentBundle()
    notes = {
        "adapter_notes": [
            "POI mirrored as EntityType.LOCATION because current Quest.location is audited as an entity reference.",
            "文本表 projected into DialogueRef(locale=...) because there is no LocalizedText model.",
        ]
    }
    p = DATA / "02_配置表" / "阵营表.xlsx"
    for r in rows(p):
        b.entities[str(r["编号"])] = Entity(
            id=str(r["编号"]),
            name=str(r["名称"]),
            type=EntityType.FACTION,
            description=str(r.get("描述") or ""),
            tags=split(r.get("标签")),
            status=str(r.get("状态") or "active"),
            source_ref=src(p, r),
        )
    p = DATA / "02_配置表" / "NPC表.xlsx"
    for r in rows(p):
        e = Entity(
            id=str(r["编号"]),
            name=str(r["名称"]),
            type=EntityType.NPC,
            description=str(r.get("描述") or ""),
            aliases=split(r.get("别名")),
            tags=split(r.get("标签")),
            status=str(r.get("状态") or "active"),
            metadata={"faction": r.get("所属阵营")},
            source_ref=src(p, r),
        )
        b.entities[e.id] = e
        if r.get("所属阵营"):
            b.relations.append(
                Relation(
                    source=e.id, target=str(r["所属阵营"]), kind="member_of", source_ref=src(p, r)
                )
            )
    p = DATA / "02_配置表" / "事件表.xlsx"
    for r in rows(p):
        order = opt_int(r.get("时间线序"))
        b.entities[str(r["编号"])] = Entity(
            id=str(r["编号"]),
            name=str(r["名称"]),
            type=EntityType.EVENT,
            description=str(r.get("描述") or ""),
            tags=[f"order={order}"] if order is not None else [],
            metadata={"timeline_order": order},
            source_ref=src(p, r),
        )
    p = DATA / "02_配置表" / "区域表.xlsx"
    for r in rows(p):
        b.regions[str(r["区域编号"])] = RegionBrief(
            id=str(r["区域编号"]),
            name=str(r["名称"]),
            level_min=opt_int(r.get("最低等级")),
            level_max=opt_int(r.get("最高等级")),
            themes=split(r.get("主题")),
            banned_content=split(r.get("禁用元素")),
            metadata={"dominant_faction": r.get("控制阵营"), "description": r.get("描述")},
            source_ref=src(p, r),
        )
    p = DATA / "02_配置表" / "POI表.xlsx"
    for r in rows(p):
        level = opt_int(r.get("等级"))
        poi = POI(
            id=str(r["编号"]),
            name=str(r["名称"]),
            region_id=str(r["所属区域"]) if r.get("所属区域") else None,
            purpose=str(r.get("叙事目的") or ""),
            controlling_faction=str(r["控制阵营"]) if r.get("控制阵营") else None,
            level_min=level,
            level_max=level,
            tags=split(r.get("标签")),
            metadata={
                "gameplay_purpose": r.get("玩法目的"),
                "linked_quests": split(r.get("关联任务")),
            },
            source_ref=src(p, r),
        )
        b.pois[poi.id] = poi
        b.entities[poi.id] = Entity(
            id=poi.id,
            name=poi.name,
            type=EntityType.LOCATION,
            description=poi.purpose,
            tags=poi.tags,
            metadata={"region_id": poi.region_id, "controlling_faction": poi.controlling_faction},
            source_ref=src(p, r),
        )
        if poi.controlling_faction:
            b.relations.append(
                Relation(
                    source=poi.id,
                    target=poi.controlling_faction,
                    kind="controlled_by",
                    source_ref=src(p, r),
                )
            )
    for q in quests_from(DATA / "02_配置表" / "任务表.xlsx"):
        b.quests[q.id] = q
    p = DATA / "02_配置表" / "任务事件引用表.xlsx"
    for r in rows(p):
        q = b.quests.get(str(r.get("任务ID")))
        if q:
            refs = list(q.metadata.get("references_event_results", []))
            refs.append(str(r.get("引用事件")))
            q.metadata["references_event_results"] = refs
    p = DATA / "02_配置表" / "对白表.xlsx"
    for r in rows(p):
        b.dialogues[str(r["对白ID"])] = DialogueRef(
            id=str(r["对白ID"]),
            text_key=str(r.get("文本Key") or r["对白ID"]),
            speaker_id=str(r["说话人"]) if r.get("说话人") else None,
            quest_id=str(r["所属任务"]) if r.get("所属任务") else None,
            text=str(r["中文文本"]) if r.get("中文文本") else None,
            locale="zh-CN",
            metadata={"vo_id": r.get("语音ID"), "ui_max_len": r.get("UI限长")},
            source_ref=src(p, r),
        )
    p = DATA / "02_配置表" / "文本表.xlsx"
    for r in rows(p):
        key = str(r["文本Key"])
        for col, locale in (("zh-CN", "zh-CN"), ("en-US", "en-US")):
            if r.get(col):
                did = f"loc_{safe_id(key)}_{safe_id(locale)}"
                b.dialogues[did] = DialogueRef(
                    id=did, text_key=key, text=str(r[col]), locale=locale, source_ref=src(p, r)
                )
    p = DATA / "02_配置表" / "术语表.xlsx"
    term_ids = {
        "真气": "term_zhenqi",
        "境界": "term_jingjie",
        "缉司": "term_jisi",
        "(敏感词)": "term_sensitive",
    }
    for i, r in enumerate(rows(p), start=1):
        canonical = str(r.get("标准词") or "")
        b.terms[term_ids.get(canonical, f"term_{i:03d}")] = Term(
            id=term_ids.get(canonical, f"term_{i:03d}"),
            canonical=canonical,
            forbidden=split(r.get("禁用词")),
            description=str(r.get("说明") or ""),
            source_ref=src(p, r),
        )
    p = DATA / "02_配置表" / "阵营关系表.xlsx"
    for r in rows(p):
        b.relations.append(
            Relation(
                source=str(r.get("来源阵营") or ""),
                target=str(r.get("目标阵营") or ""),
                kind=str(r.get("关系") or ""),
                valid_from=opt_int(r.get("生效起")),
                valid_until=opt_int(r.get("生效止")),
                metadata={"note": r.get("备注")},
                source_ref=src(p, r),
            )
        )
    md = DATA / "01_设定文档" / "雾隐山河_世界观总纲.md"
    b.style_guides["style_guide"] = StyleGuide(
        id="style_guide", body=md.read_text(encoding="utf-8"), source_ref=SourceRef(path=str(md))
    )
    notes["counts"] = counts(b)
    return b, notes


def flat(issue: Issue) -> dict[str, Any]:
    return {
        "rule_code": issue.rule_code,
        "severity": issue.severity.value,
        "category": issue.category.value,
        "target_ref": issue.target_ref,
        "message": issue.message,
        "status": issue.status.value,
        "fingerprint": issue.fingerprint or issue_fingerprint(issue),
        "evidence": [x.model_dump(mode="json", exclude_none=True) for x in issue.evidence],
    }


def has(issues: list[Issue], rule: str, target: str | None = None, text: str | None = None) -> bool:
    for issue in issues:
        blob = json.dumps(flat(issue), ensure_ascii=False)
        if (
            issue.rule_code == rule
            and (target is None or issue.target_ref == target)
            and (text is None or text in blob)
        ):
            return True
    return False


def audit_bundle(bundle: ContentBundle, baseline: AuditBaseline | None = None) -> tuple[Any, float]:
    runner = AuditRunner(build_default_rule_registry(), baseline=baseline)
    t0 = time.perf_counter()
    result = runner.run(AuditContext.from_bundle(bundle))
    return result, time.perf_counter() - t0


def seeded_eval(
    first: list[Issue], imports: list[Issue], inc: list[Issue], luban: dict[str, Any]
) -> dict[str, Any]:
    checks = {
        "E01": (
            has(first, "UNKNOWN_ENTITY_REF", "quest:q_200206", "npc_xu_sanqian"),
            "q_200206 unknown giver",
        ),
        "E02": (
            has(first, "DEPRECATED_ENTITY_REF", "quest:q_200205", "npc_yun_canzhi"),
            "deprecated NPC",
        ),
        "E03": (
            has(first, "MISSING_DIALOGUE_REF", "quest:q_200206", "dlg_200206_01"),
            "missing dialogue",
        ),
        "E04": (
            has(first, "MISSING_LOCALIZATION_KEY", "quest:q_200210"),
            "missing localization key",
        ),
        "E05": (has(first, "TEXT_TOO_LONG_FOR_UI", "dialogue:dlg_200204_01"), "UI row max length"),
        "E06": (
            has(first, "PLACEHOLDER_MISMATCH", "dialogue_key:DLG_200203_01"),
            "placeholder mismatch",
        ),
        "E07": (
            has(first, "TERM_INCONSISTENT", "dialogue:dlg_200202_01", "内力值"),
            "forbidden term 内力值",
        ),
        "E08": (
            has(first, "TERM_INCONSISTENT", "dialogue:dlg_200201_01", "氪金"),
            "forbidden term 氪金",
        ),
        "E09": (
            has(first, "MISSING_RELATION_ENDPOINT", text="fac_yanyun"),
            "missing relation endpoint",
        ),
        "E10": (
            has(first, "DUPLICATE_RELATION", "relation:fac_xuantie:enemy_of:fac_heifeng"),
            "duplicate relation",
        ),
        "E11": (has(first, "RELATION_CONFLICT", text="fac_caobang"), "relation conflict"),
        "E12": (has(first, "PREREQ_CYCLE", "quest_prerequisites"), "prereq cycle"),
        "E13": (has(first, "FACTION_CONFLICT", "quest:q_200204"), "faction conflict"),
        "E14": (has(first, "TIMELINE_VIOLATION", "quest:q_200207"), "timeline violation"),
        "E15": (
            has(first, "EVENT_RESULT_REFERENCED_TOO_EARLY", "quest:q_200208"),
            "event too early",
        ),
        "E16": (
            has(first, "CHARACTER_STATE_CONTRADICTION", "entity:npc_fang_qianli"),
            "dead+active",
        ),
        "E17": (
            has(first, "REGION_LEVEL_BOUNDS_INVALID", "region:reg_heifengling"),
            "bad region level",
        ),
        "E18": (has(first, "POI_LEVEL_OUT_OF_BOUNDS", "poi:poi_caoyun_yard"), "POI out of bounds"),
        "E19": (
            has(first, "POI_WITHOUT_NARRATIVE_PURPOSE", "poi:poi_shuiyue_temple"),
            "empty POI purpose",
        ),
        "E20": (has(first, "REGION_BANNED_CONTENT_USED", "poi:poi_xunshan_camp"), "banned content"),
        "E21": (has(first, "QUEST_MISSING_OBJECTIVE", "quest:q_200209"), "empty objective"),
        "E22": (has(imports, "IMPORT_CONFLICT", "quest:q_200201"), "import conflict"),
        "E23": (
            has(inc, "PREREQ_MISSING", "quest:q_200303")
            or has(inc, "UNKNOWN_ENTITY_REF", "quest:q_200303", "q_200399"),
            "missing prereq probe",
        ),
        "E24": (
            any("INJECTION" in x.rule_code or "PROMPT" in x.rule_code for x in first),
            "prompt injection probe",
        ),
        "E25": (bool(luban.get("supported")), "luban header probe"),
        "E26": (
            has(inc, "UNKNOWN_ENTITY_REF", "quest:q_200302", "npc_bai_su"),
            "increment typo NPC",
        ),
    }
    details = {k: {"detected": v[0], "note": v[1]} for k, v in checks.items()}
    surface = [f"E{i:02d}" for i in range(1, 23)] + ["E26"]
    return {
        "details": details,
        "rule_surface_detected": sum(1 for x in surface if details[x]["detected"]),
        "rule_surface_total": len(surface),
        "rule_surface_rate": round(
            sum(1 for x in surface if details[x]["detected"]) / len(surface), 4
        ),
        "probe_detected_or_supported": sum(
            1 for x in ["E23", "E24", "E25"] if details[x]["detected"]
        ),
        "probe_total": 3,
    }


def stage2_qa() -> dict[str, Any]:
    results = []
    for i, q in enumerate(read_json(DATA / "04_答案卷" / "qa_questions.json"), start=1):
        query = q["q"]
        name = f"q{i:02d}_{safe_id(query)[:20]}"
        c = run_cli(
            f"stage2_context_{i:02d}",
            ["context-pack", "--content-root", str(CONTENT), "--query", query],
            RUN / "stage2_qa" / f"{name}_context.json",
        )
        a = run_cli(
            f"stage2_ask_{i:02d}",
            ["ask", "--content-root", str(CONTENT), "--query", query, "--max-cost-usd", "0.01"],
            RUN / "stage2_qa" / f"{name}_ask.json",
        )
        cp = c.get("parsed") or {}
        ap = a.get("parsed") or {}
        ans = ap.get("answer") or {}
        refs = set(cp.get("refs") or [])
        citations = [x.get("ref") for x in ans.get("citations", []) if isinstance(x, dict)]
        expect_terms = [x for x in re.split(r"[;(),，。/\s]+", str(q.get("expect") or "")) if x]
        results.append(
            {
                "index": i,
                "question": query,
                "type": q.get("type"),
                "expect": q.get("expect"),
                "expected_refuse": q.get("refuse"),
                "context_hit_count": len(cp.get("hits") or []),
                "context_refs": sorted(refs),
                "answer": ans,
                "citations_in_context": all(x in refs for x in citations),
                "refusal_correct": bool(ans.get("refused")) == bool(q.get("refuse")),
                "semantic_contains_expected_term": any(
                    x in str(ans.get("answer") or "") for x in expect_terms[:4]
                ),
                "ask_duration_sec": a["duration_sec"],
            }
        )
    out = {
        "questions": len(results),
        "context_non_empty": sum(x["context_hit_count"] > 0 for x in results),
        "citation_valid": sum(x["citations_in_context"] for x in results),
        "refusal_correct": sum(x["refusal_correct"] for x in results),
        "semantic_contains_expected_term": sum(
            x["semantic_contains_expected_term"] for x in results
        ),
        "mean_ask_duration_sec": round(
            sum(x["ask_duration_sec"] for x in results) / len(results), 4
        ),
        "results": results,
        "observation": "CLI ask uses OfflineQAProvider; answers are generic cited smoke-test outputs, not factual Chinese answers.",
    }
    write_json(RUN / "stage2_qa_results.json", out)
    return out


def stage3_increment(baseline: AuditBaseline) -> tuple[dict[str, Any], list[Issue], list[Issue]]:
    path = DATA / "03_新版本提交" / "支线_雾隐走私案_v0610.xlsx"
    mapping = FieldMapping(
        default_kind="quest",
        columns={
            "任务ID": "id",
            "任务名": "title",
            "接取NPC": "giver_npc",
            "发生地点": "location",
            "前置任务": "prerequisites",
            "任务目标": "objective",
            "对白引用": "dialogue_refs",
            "文本Key": "localization_keys",
        },
    )
    t0 = time.perf_counter()
    dry = ingest_paths([path], store=ContentStore(CONTENT), dry_run=True, field_mapping=mapping)
    ingest_sec = time.perf_counter() - t0
    store = ContentStore(CONTENT)
    bundle = store.load()
    applied, skipped = [], []
    for q in quests_from(path):
        if q.id == "q_200201":
            skipped.append(q.id)
        else:
            bundle.quests[q.id] = q
            applied.append(q.id)
    store.save(bundle)
    audit, audit_sec = audit_bundle(bundle, baseline=baseline)
    gate = run_cli(
        "stage3_cli_audit_gate",
        ["audit", "--content-root", str(CONTENT), "--baseline", str(BASELINE), "--fail-on-error"],
        RUN / "stage3_cli_audit_gate.json",
    )
    out = {
        "duration_sec_ingest": round(ingest_sec, 4),
        "duration_sec_audit": round(audit_sec, 4),
        "applied_increment_quests": applied,
        "skipped_conflict_quests": skipped,
        "dry_run_import_issues": [flat(x) for x in dry.issues],
        "audit_open_errors": len(audit.open_errors),
        "audit_issues": [flat(x) for x in audit.issues],
        "cli_gate_exit_code": gate["exit_code"],
    }
    write_json(RUN / "stage3_increment_results.json", out)
    return out, dry.issues, audit.issues


def stage4_patch() -> dict[str, Any]:
    store = ContentStore(CONTENT)
    before = store.load()
    before_hash = content_hash(before)
    runner = AuditRunner(build_default_rule_registry())
    raw = json.dumps(
        {
            "candidates": [
                {
                    "ops": [
                        {
                            "op": "add",
                            "path": "/quests/q_200210/localization_keys/-",
                            "value": "QUEST_200210_NAME",
                        },
                        {
                            "op": "replace",
                            "path": "/regions/reg_heifengling/level_min",
                            "value": 18,
                        },
                    ],
                    "rationale": "Fix E04/E17",
                },
                {
                    "ops": [
                        {
                            "op": "replace",
                            "path": "/quests/q_200210/giver_npc",
                            "value": "npc_missing_patch",
                        }
                    ],
                    "rationale": "Bad candidate",
                },
            ]
        },
        ensure_ascii=False,
    )
    candidates = parse_patch_candidates(raw)
    validations = [validate_patch_candidate(before, c, runner) for c in candidates]
    valid = valid_patch_candidates(before, candidates, runner)
    applied = apply_patch_to_store(store, valid[0].candidate, applied_by="wuyin_experiment")
    after_apply = store.load()
    after_audit, _ = audit_bundle(after_apply)
    rollback_error = None
    native_rollback_restored = False
    try:
        rollback_patch_in_store(store, applied.rollback_ops)
        after_rollback = store.load()
        native_rollback_restored = before_hash == content_hash(after_rollback)
    except Exception as exc:  # Keep the rollback defect as an experiment result, then restore.
        rollback_error = f"{exc.__class__.__name__}: {exc}"
        store.save(before)
        after_rollback = store.load()
    rollback_audit, _ = audit_bundle(after_rollback)
    out = {
        "before_hash": before_hash,
        "after_apply_hash": content_hash(after_apply),
        "after_rollback_hash": content_hash(after_rollback),
        "native_rollback_restored_content_hash": native_rollback_restored,
        "rollback_error": rollback_error,
        "manual_restore_after_rollback_error": rollback_error is not None,
        "rollback_restored_content_hash": before_hash == content_hash(after_rollback),
        "candidate_count": len(candidates),
        "valid_candidate_count": len(valid),
        "validations": [
            {
                "valid": v.valid,
                "introduced_errors": v.introduced_errors,
                "resolved_errors": v.resolved_errors,
                "candidate": v.candidate,
            }
            for v in validations
        ],
        "applied_candidate_status": applied.candidate.status.value,
        "issue_presence_after_apply": {
            "E04": has(after_audit.issues, "MISSING_LOCALIZATION_KEY", "quest:q_200210"),
            "E17": has(after_audit.issues, "REGION_LEVEL_BOUNDS_INVALID", "region:reg_heifengling"),
        },
        "issue_presence_after_rollback": {
            "E04": has(rollback_audit.issues, "MISSING_LOCALIZATION_KEY", "quest:q_200210"),
            "E17": has(
                rollback_audit.issues, "REGION_LEVEL_BOUNDS_INVALID", "region:reg_heifengling"
            ),
        },
    }
    write_json(RUN / "stage4_patch_results.json", out)
    return out


def stage5_impact() -> dict[str, Any]:
    project = ProjectContext.open(CONTENT, sqlite_path=RUN / "stage5.sqlite")
    try:
        analyzer = ImpactAnalyzer(project.graph)
        scenarios = {
            "S1_delete_lu_wang": ChangeSet(
                changes=[
                    Change(change_type=ChangeType.ENTITY_DELETE, target_ref="entity:npc_lu_wang")
                ]
            ),
            "S2_change_canglang_faction": ChangeSet(
                changes=[
                    Change(change_type=ChangeType.RELATION_CHANGE, target_ref="entity:fac_canglang")
                ]
            ),
            "S3_rename_poi_node": ChangeSet(
                changes=[
                    Change(change_type=ChangeType.ENTITY_RENAME, target_ref="poi:poi_wuyin_dock")
                ]
            ),
            "S3_rename_poi_entity_mirror": ChangeSet(
                changes=[
                    Change(change_type=ChangeType.ENTITY_RENAME, target_ref="entity:poi_wuyin_dock")
                ]
            ),
        }
        results = {k: analyzer.analyze(v).model_dump(mode="json") for k, v in scenarios.items()}
    finally:
        project.close()
    expected = {
        "quest:q_100101",
        "quest:q_100102",
        "quest:q_200204",
        "quest:q_200210",
        "dialogue:dlg_100101_01",
        "dialogue:dlg_200204_01",
    }
    actual = {
        x["target_ref"]
        for x in results["S1_delete_lu_wang"]["items"]
        if x["level"] == "must_change"
    }
    out = {
        "results": results,
        "recall_checks": {
            "S1_must_expected": sorted(expected),
            "S1_must_actual": sorted(actual),
            "S1_must_recall": round(len(expected & actual) / len(expected), 4),
            "S3_poi_node_actual_must": sorted(
                x["target_ref"]
                for x in results["S3_rename_poi_node"]["items"]
                if x["level"] == "must_change"
            ),
            "S3_entity_mirror_actual_must": sorted(
                x["target_ref"]
                for x in results["S3_rename_poi_entity_mirror"]["items"]
                if x["level"] == "must_change"
            ),
        },
        "observation": "Relation edges are not first-class change targets; quest.location currently reaches entity:poi_* mirror rather than poi:* node.",
    }
    write_json(RUN / "stage5_impact_results.json", out)
    return out


def stage6_gate() -> dict[str, Any]:
    store = ContentStore(CONTENT)
    fix = apply_patch_to_store(
        store,
        PatchCandidate(
            ops=[
                PatchOperation(
                    op=PatchOp.REPLACE, path="/quests/q_200302/giver_npc", value="npc_bai_susu"
                )
            ]
        ),
        applied_by="wuyin_experiment",
    )
    green = run_cli(
        "stage6_gate_after_e26_fix",
        ["audit", "--content-root", str(CONTENT), "--baseline", str(BASELINE), "--fail-on-error"],
        RUN / "stage6_gate_after_e26_fix.json",
    )
    broken = apply_patch_to_store(
        store,
        PatchCandidate(
            ops=[
                PatchOperation(
                    op=PatchOp.REPLACE, path="/quests/q_200301/giver_npc", value="npc_missing_gate"
                )
            ]
        ),
        applied_by="wuyin_experiment",
    )
    red = run_cli(
        "stage6_gate_after_break",
        ["audit", "--content-root", str(CONTENT), "--baseline", str(BASELINE), "--fail-on-error"],
        RUN / "stage6_gate_after_break.json",
    )
    rollback_patch_in_store(store, broken.rollback_ops)
    green2 = run_cli(
        "stage6_gate_after_restore",
        ["audit", "--content-root", str(CONTENT), "--baseline", str(BASELINE), "--fail-on-error"],
        RUN / "stage6_gate_after_restore.json",
    )
    out = {
        "fix_e26_status": fix.candidate.status.value,
        "after_fix_exit_code": green["exit_code"],
        "after_break_exit_code": red["exit_code"],
        "after_restore_exit_code": green2["exit_code"],
        "red_green_deterministic": green["exit_code"] == 0
        and red["exit_code"] == 1
        and green2["exit_code"] == 0,
    }
    write_json(RUN / "stage6_gate_results.json", out)
    return out


class FixedProvider:
    def __init__(self, payload: Any, in_tok: int = 20, out_tok: int = 10) -> None:
        self.payload = payload
        self.in_tok = in_tok
        self.out_tok = out_tok

    def complete(self, *, system: str, user: str, model: str) -> tuple[str, int, int]:
        return json.dumps(self.payload, ensure_ascii=False), self.in_tok, self.out_tok


def stage7_generation() -> dict[str, Any]:
    project = ProjectContext.open(CONTENT, sqlite_path=RUN / "stage7.sqlite")
    try:
        dtel = TelemetryCollector()
        draft_gateway = LLMGateway(
            providers={
                "cheap": FixedProvider(
                    {
                        "id": "q_ai_wuyin_infiltration",
                        "title": "雾隐潜声",
                        "giver_npc": "npc_bai_susu",
                        "location": "poi_wuyin_dock",
                        "objective": "潜入雾隐渡口,探查私盐快船的出航暗号。",
                        "prerequisites": ["q_100102"],
                        "timeline_order": 23,
                        "localization_keys": ["QUEST_AI_WUYIN_INFILTRATION_NAME"],
                        "dialogue_refs": [],
                        "tags": ["side", "infiltration"],
                    },
                    120,
                    80,
                )
            },
            router=StaticRouter(mapping={"quest_draft": "cheap"}),
            cache=NoOpCache(),
            telemetry=dtel,
        )
        draft = QuestDraftService(
            gateway=draft_gateway,
            context_builder=project.context_builder,
            audit_runner=AuditRunner(build_default_rule_registry()),
            bundle=project.bundle,
        ).draft_quest("雾隐泽潜入支线")
        btel = TelemetryCollector()
        bark_gateway = LLMGateway(
            providers={
                "cheap": FixedProvider(
                    {
                        "variants": [
                            "站住,报上来路。",
                            "雾里别乱走。",
                            "陆忘的人也敢来?",
                            "想靠氪金买命?",
                            "这句台词故意写得非常非常非常非常长超过限制",
                            "把船灯压低。",
                            "刀收好,别惊了水。",
                            "白素素在此。",
                        ]
                    },
                    80,
                    60,
                )
            },
            router=StaticRouter(mapping={"barks_batch": "cheap"}),
            cache=NoOpCache(),
            telemetry=btel,
        )
        queue = ReviewQueue()
        barks = BarkBatchService(
            gateway=bark_gateway, bundle=project.bundle, review_queue=queue
        ).generate(
            speaker_ids=["npc_bai_susu"],
            topic="发现入侵者",
            variants_per_speaker=8,
            max_chars=30,
            allowed_entity_ids={"npc_bai_susu"},
        )
        if queue.list_pending():
            queue.mark(queue.list_pending()[0].id, "approved")
    finally:
        project.close()
    out = {
        "draft": draft.model_dump(mode="json"),
        "draft_issue_codes": sorted({x.rule_code for x in draft.issues}),
        "draft_origin": draft.quest.origin.value,
        "draft_review_status": draft.quest.review_status.value,
        "draft_telemetry": dtel.summary(),
        "bark_accept_count": len(barks.accepted),
        "bark_reject_count": len(barks.rejected),
        "bark_reject_codes": [i.code for r in barks.rejected for i in r.issues],
        "barks": barks.model_dump(mode="json"),
        "review_queue_pending_after_mark_one": len(queue.list_pending()),
        "bark_telemetry": btel.summary(),
    }
    write_json(RUN / "stage7_generation_results.json", out)
    return out


def stage8_export() -> dict[str, Any]:
    export = run_cli(
        "stage8_export_unity",
        [
            "export",
            "--content-root",
            str(CONTENT),
            "--output-dir",
            str(EXPORTS),
            "--target-engine",
            "unity",
        ],
        RUN / "stage8_export_cli.json",
    )
    manifest_path = EXPORTS / "unity" / "manifest.json"
    bundle_path = EXPORTS / "unity" / "content_bundle.json"
    manifest = read_json(manifest_path)
    file_sha = hashlib.sha256(bundle_path.read_bytes()).hexdigest()
    recorded = manifest["files"][0]["sha256"] if manifest.get("files") else None
    golden = run_cli(
        "stage8_eval_golden",
        ["eval-golden", "--workspace", str(RUN / "golden_workspace")],
        RUN / "stage8_eval_golden.json",
    )
    out = {
        "export_exit_code": export["exit_code"],
        "manifest_path": str(manifest_path),
        "bundle_path": str(bundle_path),
        "manifest": manifest,
        "content_hash_now": content_hash(ContentStore(CONTENT).load()),
        "manifest_content_hash_matches_current": manifest.get("content_hash")
        == content_hash(ContentStore(CONTENT).load()),
        "manifest_recorded_sha": recorded,
        "actual_bundle_file_sha256": file_sha,
        "manifest_file_sha_matches_bytes": recorded == file_sha,
        "eval_golden_exit_code": golden["exit_code"],
        "eval_golden": golden.get("parsed"),
        "observation": "Exporter writes engine-agnostic JSON; manifest file sha is canonical payload hash, not literal file bytes.",
    }
    write_json(RUN / "stage8_export_results.json", out)
    return out


def write_report(summary: dict[str, Any]) -> None:
    seeded = summary["seeded_error_evaluation"]
    lines = [
        "# 雾隐山河真实场景实验结果记录",
        "",
        f"- 执行时间: {datetime.now().isoformat(timespec='seconds')}",
        f"- 数据包: `{DATA}`",
        f"- 内容仓: `{CONTENT}`",
        f"- 报告目录: `{RUN}`",
        f"- 导出目录: `{EXPORTS}`",
        "",
        "## 总览",
        "",
        f"- 适配器导入对象数: {summary['stage1']['adapter_notes']['counts']}",
        f"- 首次审计: issue {summary['stage1']['audit_first']['issue_count']} 条, open error {summary['stage1']['audit_first']['open_errors']} 条, 耗时 {summary['stage1']['audit_first']['duration_sec']} 秒。",
        f"- 规则面埋错检出: {seeded['rule_surface_detected']}/{seeded['rule_surface_total']} = {seeded['rule_surface_rate']}。",
        f"- 缺口探测器命中/支持: {seeded['probe_detected_or_supported']}/{seeded['probe_total']}。",
        f"- QA: context 非空 {summary['stage2_qa']['context_non_empty']}/{summary['stage2_qa']['questions']}, citation 有效 {summary['stage2_qa']['citation_valid']}/{summary['stage2_qa']['questions']}, 拒答正确 {summary['stage2_qa']['refusal_correct']}/{summary['stage2_qa']['questions']}, 语义命中 {summary['stage2_qa']['semantic_contains_expected_term']}/{summary['stage2_qa']['questions']}。",
        f"- 增量门禁: q_200201 冲突检出={seeded['details']['E22']['detected']}, E26 检出={seeded['details']['E26']['detected']}, CLI exit={summary['stage3_increment']['cli_gate_exit_code']}。",
        f"- Patch 回滚: content_hash 还原={summary['stage4_patch']['rollback_restored_content_hash']}, valid candidates={summary['stage4_patch']['valid_candidate_count']}/{summary['stage4_patch']['candidate_count']}。",
        f"- 影响分析 S1 召回: {summary['stage5_impact']['recall_checks']['S1_must_recall']}。",
        f"- 锁表门禁红绿翻转: {summary['stage6_gate']['red_green_deterministic']}。",
        f"- 生成: draft issues={summary['stage7_generation']['draft_issue_codes']}, barks accepted/rejected={summary['stage7_generation']['bark_accept_count']}/{summary['stage7_generation']['bark_reject_count']}。",
        f"- 导出: content_hash匹配={summary['stage8_export']['manifest_content_hash_matches_current']}, 文件字节sha匹配={summary['stage8_export']['manifest_file_sha_matches_bytes']}, eval-golden exit={summary['stage8_export']['eval_golden_exit_code']}。",
        "",
        "## 幕一 冷启动接入",
        "",
        "- CLI 无 `--field-mapping`; 原始中文 XLSX dry-run 不能形成可用业务对象。",
        "- Markdown 兼容解析不能正确处理本数据包中的 `(id)` 条目和 `source kind target` 关系写法。",
        "- Luban 表头探测失败: `##var/##type/##group` 被当成普通表头/数据。",
        "- 采用不改源码的实验适配器后，下游审计可运行；适配器明确标记了当前模型缺口。",
        "",
        "### 埋错命中",
        "",
    ]
    for eid, item in seeded["details"].items():
        lines.append(f"- {eid}: {'✓' if item['detected'] else '✗'} {item['note']}")
    lines += [
        "",
        "### 额外告警",
        "",
    ]
    unexpected = summary["stage1"]["audit_first"]["unexpected_issues"]
    if unexpected:
        for item in unexpected:
            lines.append(f"- {item['rule_code']} {item['target_ref']}: {item['message']}")
    else:
        lines.append("- 无。")
    lines += [
        "",
        "## 幕二 日常查设定",
        "",
        "- 当前 `ask` 是离线 smoke test provider，返回泛化引用句，不生成事实答案；语义准确率因此失败。",
    ]
    for item in summary["stage2_qa"]["results"]:
        lines.append(
            f"- Q{item['index']:02d} {item['type']}: hits={item['context_hit_count']}, refused={item['answer'].get('refused')}, citation_ok={item['citations_in_context']}, semantic_hit={item['semantic_contains_expected_term']}"
        )
    lines += [
        "",
        "## 幕三 周版本增量提交",
        "",
        f"- 已应用非冲突任务: {summary['stage3_increment']['applied_increment_quests']}; 跳过冲突任务: {summary['stage3_increment']['skipped_conflict_quests']}。",
        f"- baseline 后新增 open error: {summary['stage3_increment']['audit_open_errors']}; CLI gate exit: {summary['stage3_increment']['cli_gate_exit_code']}。",
        "- E23 前置缺失未被现有规则接住，应新增 `PREREQ_MISSING` 或把 prerequisites 纳入引用完整性检查。",
        "",
        "## 幕四 修复闭环",
        "",
        f"- 候选数 {summary['stage4_patch']['candidate_count']}, 通过安全校验 {summary['stage4_patch']['valid_candidate_count']}。坏候选因引入新 error 被过滤。",
        f"- Apply 后 E04/E17 是否仍存在: {summary['stage4_patch']['issue_presence_after_apply']}。",
        f"- Rollback 后 E04/E17 是否恢复: {summary['stage4_patch']['issue_presence_after_rollback']}。",
        f"- 原生 rollback 是否逐 hash 还原: {summary['stage4_patch']['native_rollback_restored_content_hash']}; rollback_error={summary['stage4_patch']['rollback_error']}; 已手动恢复继续后续实验={summary['stage4_patch']['manual_restore_after_rollback_error']}。",
        "",
        "## 幕五 影响分析",
        "",
        f"- S1 must_change 实际: {summary['stage5_impact']['recall_checks']['S1_must_actual']}。",
        "- S2 关系翻转无法以 relation edge 为直接 target，只能退化为 faction entity 变更。",
        f"- S3 `poi:*` 直接影响: {summary['stage5_impact']['recall_checks']['S3_poi_node_actual_must']}。",
        f"- S3 `entity:poi_*` 镜像直接影响: {summary['stage5_impact']['recall_checks']['S3_entity_mirror_actual_must']}。",
        "",
        "## 幕六 锁表周回归",
        "",
        f"- 修复 E26 后 exit={summary['stage6_gate']['after_fix_exit_code']}; 人为改坏 exit={summary['stage6_gate']['after_break_exit_code']}; 回滚后 exit={summary['stage6_gate']['after_restore_exit_code']}。",
        "",
        "## 幕七 受约束生成",
        "",
        f"- Quest draft origin={summary['stage7_generation']['draft_origin']}, review_status={summary['stage7_generation']['draft_review_status']}, issue_codes={summary['stage7_generation']['draft_issue_codes']}。",
        "- DraftService 当前对“整个内容包 + draft”运行审计，未套用 baseline/target 过滤，因此返回了既有存量 issue；生成内容自身的新增信任告警是 `UNREVIEWED_AI_CONTENT`。",
        f"- Bark accepted={summary['stage7_generation']['bark_accept_count']}, rejected={summary['stage7_generation']['bark_reject_count']}, reject_codes={summary['stage7_generation']['bark_reject_codes']}。",
        "",
        "## 幕八 出包交付",
        "",
        f"- Unity bundle: `{summary['stage8_export']['bundle_path']}`。",
        f"- Manifest: `{summary['stage8_export']['manifest_path']}`。",
        "- 当前导出是引擎无关 JSON bundle，不是 UE DataTable/CSV 行级导出。",
        "- manifest.files[0].sha256 与实际文件字节 sha 不一致，字段语义应修正。",
        "",
        "## Backlog",
        "",
        "- P0: CLI 支持 `--field-mapping` 与按文件 default_kind。",
        "- P0: XLSX importer 支持 luban 多行表头或明确报 unsupported。",
        "- P0: 新增 `PREREQ_MISSING`。",
        "- P0: `TEXT_TOO_LONG_FOR_UI` 读取行级 `ui_max_len`。",
        "- P1: 将 NPC 阵营、POI 控制、任务事件引用、多语言文本建成一等模型/派生边。",
        "- P1: v2 ingest 恢复 prompt-injection 扫描。",
        "- P1: `ask` 离线路径改为真实抽取式回答或明确标记 retrieval smoke test。",
        "- P1: DraftService 审计结果支持 baseline 或只返回 draft 相关 issue，避免污染 review 队列。",
        "- P1: ImpactAnalyzer 增加关系边 target、`owcopilot impact` CLI，并统一 quest.location 到 `poi:*` 节点。",
        "- P2: Export manifest 的 `sha256` 改为文件字节 hash，或改名为 `content_payload_hash`。",
        "- P2: 补 UE DataTable/CSV 行级 exporter。",
        "",
        "## 证据文件",
        "",
    ]
    for path in sorted(RUN.glob("*.json")):
        lines.append(f"- `{path}`")
    lines.append(f"- 命令日志目录: `{RUN / 'commands'}`")
    text = "\n".join(lines) + "\n"
    (RUN / "实验结果记录.md").write_text(text, encoding="utf-8")
    (REPORTS / "实验结果记录_latest.md").write_text(text, encoding="utf-8")


def main() -> int:
    REPORTS.mkdir(parents=True, exist_ok=True)
    RUN.mkdir(parents=True, exist_ok=True)
    reset_dir(CONTENT)
    reset_dir(EXPORTS)
    write_json(
        RUN / "environment.json",
        {
            "python": sys.version,
            "root": ROOT,
            "data": DATA,
            "content": CONTENT,
            "reports": REPORTS,
            "exports": EXPORTS,
        },
    )

    run_cli(
        "stage1_cli_md_dryrun",
        [
            "ingest",
            "--content-root",
            str(CONTENT),
            "--input",
            str(DATA / "01_设定文档" / "雾隐山河_世界观总纲.md"),
        ],
        RUN / "stage1_cli_md_dryrun.json",
    )
    run_cli(
        "stage1_cli_xlsx_raw_dryrun",
        [
            "ingest",
            "--content-root",
            str(CONTENT),
            "--input",
            str(DATA / "02_配置表" / "阵营表.xlsx"),
        ],
        RUN / "stage1_cli_xlsx_raw_dryrun.json",
    )
    luban_raw = parse_paths([DATA / "02_配置表" / "NPC表_luban风格.xlsx"])
    luban_norm = normalize_raw_objects(luban_raw)
    luban = {
        "raw_count": len(luban_raw),
        "raw_preview": [x.model_dump(mode="json") for x in luban_raw[:8]],
        "normalized_counts": counts(luban_norm),
        "normalized_entity_ids": sorted(luban_norm.entities),
        "supported": sorted(luban_norm.entities) == ["npc_lu_wang", "npc_shen_qinghe"],
    }
    write_json(RUN / "stage1_luban_probe.json", luban)

    bundle, notes = build_bundle()
    ContentStore(CONTENT).save(bundle)
    write_json(RUN / "stage1_adapter_bundle_counts.json", notes)
    t0 = time.perf_counter()
    first, first_sec = audit_bundle(bundle)
    first_sec = time.perf_counter() - t0
    expected_targets = {
        ("UNKNOWN_ENTITY_REF", "quest:q_200206"),
        ("DEPRECATED_ENTITY_REF", "quest:q_200205"),
        ("MISSING_DIALOGUE_REF", "quest:q_200206"),
        ("MISSING_LOCALIZATION_KEY", "quest:q_200210"),
        ("TEXT_TOO_LONG_FOR_UI", "dialogue:dlg_200204_01"),
        ("PLACEHOLDER_MISMATCH", "dialogue_key:DLG_200203_01"),
        ("TERM_INCONSISTENT", "dialogue:dlg_200202_01"),
        ("TERM_INCONSISTENT", "dialogue:dlg_200201_01"),
        ("MISSING_RELATION_ENDPOINT", "relation:fac_yanyun:enemy_of:fac_xuantie"),
        ("DUPLICATE_RELATION", "relation:fac_xuantie:enemy_of:fac_heifeng"),
        ("RELATION_CONFLICT", "relation:fac_canglang:conflict:fac_caobang"),
        ("PREREQ_CYCLE", "quest_prerequisites"),
        ("FACTION_CONFLICT", "quest:q_200204"),
        ("TIMELINE_VIOLATION", "quest:q_200207"),
        ("EVENT_RESULT_REFERENCED_TOO_EARLY", "quest:q_200208"),
        ("CHARACTER_STATE_CONTRADICTION", "entity:npc_fang_qianli"),
        ("REGION_LEVEL_BOUNDS_INVALID", "region:reg_heifengling"),
        ("POI_LEVEL_OUT_OF_BOUNDS", "poi:poi_caoyun_yard"),
        ("POI_WITHOUT_NARRATIVE_PURPOSE", "poi:poi_shuiyue_temple"),
        ("REGION_BANNED_CONTENT_USED", "poi:poi_xunshan_camp"),
        ("QUEST_MISSING_OBJECTIVE", "quest:q_200209"),
    }
    first_payload = {
        "duration_sec": round(first_sec, 4),
        "content_hash": content_hash(bundle),
        "issue_count": len(first.issues),
        "open_errors": len(first.open_errors),
        "totals": first.run.totals,
        "baseline_delta": first.run.baseline_delta,
        "issues": [flat(x) for x in first.issues],
        "unexpected_issues": [
            flat(x) for x in first.issues if (x.rule_code, x.target_ref) not in expected_targets
        ],
    }
    write_json(RUN / "stage1_audit_first.json", first_payload)

    baseline = AuditBaseline()
    for issue in first.issues:
        baseline.add(issue)
    BASELINE.parent.mkdir(parents=True, exist_ok=True)
    write_json(BASELINE, baseline)
    write_json(RUN / "stage1_baseline.json", baseline)
    bl, bl_sec = audit_bundle(bundle, baseline)
    bl_payload = {
        "duration_sec": round(bl_sec, 4),
        "open_errors": len(bl.open_errors),
        "totals": bl.run.totals,
        "baseline_delta": bl.run.baseline_delta,
        "issues": [flat(x) for x in bl.issues],
    }
    write_json(RUN / "stage1_audit_with_baseline.json", bl_payload)

    qa = stage2_qa()
    stage3, import_issues, inc_issues = stage3_increment(baseline)
    stage4 = stage4_patch()
    stage5 = stage5_impact()
    stage6 = stage6_gate()
    stage7 = stage7_generation()
    stage8 = stage8_export()
    seeds = seeded_eval(first.issues, import_issues, inc_issues, luban)
    write_json(RUN / "seeded_error_evaluation.json", seeds)

    summary = {
        "stage1": {
            "adapter_notes": notes,
            "audit_first": first_payload,
            "audit_with_baseline": bl_payload,
            "luban_probe": luban,
        },
        "stage2_qa": qa,
        "stage3_increment": stage3,
        "stage4_patch": stage4,
        "stage5_impact": stage5,
        "stage6_gate": stage6,
        "stage7_generation": stage7,
        "stage8_export": stage8,
        "seeded_error_evaluation": seeds,
        "command_log_count": len(COMMANDS),
        "run_dir": str(RUN),
    }
    write_json(RUN / "experiment_summary.json", summary)
    write_json(REPORTS / "experiment_summary_latest.json", summary)
    write_json(RUN / "command_log.json", COMMANDS)
    write_report(summary)
    print(
        json.dumps(
            {
                "run_dir": str(RUN),
                "report": str(RUN / "实验结果记录.md"),
                "summary": str(RUN / "experiment_summary.json"),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
